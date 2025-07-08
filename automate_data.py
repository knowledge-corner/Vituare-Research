'''
Ex. Automate an Excel sheet using Python to enhance a dataset with calculated columns
- Adds 4 new columns: Rating, Bonus, Tax, Net Salary.
Applies data validation to Rating (allowed values: 1–5).
Calculates Bonus as a percentage of Salary based on Rating.
Calculates Tax as (Salary + Bonus) * 10%.
Calculates Net Salary as Salary + Bonus - Tax.
'''

import openpyxl as px
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection

import os

os.chdir(r"C:\Users\vaide\OneDrive - knowledgecorner.in\Course Material\Clients\Virtua Search\Vituare-Research\Files")

wb = px.load_workbook('input.xlsx')
ws = wb.active

new_cols = ['Rating', 'Bonus', 'Tax', 'Net Salary']
exsisting_cols = [cell.value for cell in ws[1]]
print(f"Existing columns: {exsisting_cols}")

for col in new_cols:
    if col not in exsisting_cols:
        ws.cell(row=1, column=ws.max_column + 1, value=col)

headers = [cell.value for cell in ws[1]]
col_letters = {name : get_column_letter(index) for index, name in enumerate(headers, start = 1)}

# Adding data validation for Rating
rating_col = col_letters['Rating']
# rating_range = f"{rating_col}2:{rating_col}{ws.max_row}"  # D2:D3
rating_range = f"{rating_col}2:{rating_col}1048576"  # till the end of the column
print(f"Rating range: {rating_range}")

dv = DataValidation(type="list", formula1 = '"Excellent,Good,Average,Below Average,Poor"', allow_blank=True)
dv.prompt = "Select a rating"
dv.error = "Invalid rating. Please select from the list."
ws.add_data_validation(dv)
dv.add(rating_range)

# Calculating Bonus, Tax, and Net Salary
for row in range(2, ws.max_row + 1):
    salary = f"{col_letters['Salary']}{row}"
    rating = f"{col_letters['Rating']}{row}"
    bonus = f"{col_letters['Bonus']}{row}"
    tax = f"{col_letters['Tax']}{row}"
    net = f"{col_letters['Net Salary']}{row}"

    ws[bonus] = f"= IF({rating} = \"Excellent\", {salary} * 0.12, IF({rating} = \"Good\", {salary} * 0.10, IF({rating} = \"Average\", {salary} * 0.08, IF({rating} = \"Below Average\", {salary} * 0.05, IF({rating} = \"Poor\", {salary} * 0.02, 0)))))"
    ws[tax] = f"=({salary} + {bonus}) * 0.1"
    ws[net] = f"={salary} + {bonus} - {tax}"
    
ws.column_dimensions[col_letters["Bonus"]].hidden = True
ws.column_dimensions[col_letters["Tax"]].hidden = True

rating_col_index = headers.index('Rating') + 1
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column): 
    for cell in row:
        if cell.column == rating_col_index or row[0].row > 1:  # Skip the header row and Rating column
            cell.protection = Protection(locked=False, hidden=True)
        else:
            cell.protection = Protection(locked=True, hidden=True)

ws.protection.enable()     
ws.protection.set_password('password')  # Set a password for the sheet protection  
         
wb.save('input.xlsx')
