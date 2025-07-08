import shutil
import openpyxl as px
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

import os

os.chdir(r"C:\Users\vaide\OneDrive - knowledgecorner.in\Course Material\Clients\Virtua Search\Vituare-Research\Files")

shutil.copy("input.xlsx", "destination.xlsx")

# Apply conditional formating to rows where Rating is "Poor"

wb = px.load_workbook("destination.xlsx")
ws = wb.active

fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

existing_cols = [cell.value for cell in ws[1]]
rating_index = get_column_letter(existing_cols.index('Rating') + 1)
max_column_letter = get_column_letter(ws.max_column)

for row in range(2, ws.max_row + 1):
    formula = f"=${rating_index}{row} = \"Poor\""
    ws.conditional_formatting.add(f"A{row}:{max_column_letter}{row}", FormulaRule([formula], fill=fill))

wb.save("destination.xlsx")
# wb.close()